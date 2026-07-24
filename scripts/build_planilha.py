#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera planilha-q22-l8.xlsx a partir de dados/*.json.
Rodado automaticamente pelo GitHub Actions a cada alteração em dados/ (cotações
lançadas pelo site, avanço de etapas). NÃO editar a planilha à mão — ela é
sobrescrita a cada geração. A fonte de dados é o site/os JSONs.
"""
import json, datetime, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = pathlib.Path(__file__).resolve().parent.parent
DADOS = RAIZ / 'dados'
SAIDA = RAIZ / 'planilha-q22-l8.xlsx'

AZUL = '0E3A6B'; AZUL2 = '2E5E96'; CINZA = 'EEF4FB'; AMBAR = 'B8860B'; VERDE = '1E7A46'
TIPO = {'material': 'material', 'mao_obra': 'mão de obra', 'servico': 'serviço'}
RS = '"R$" #,##0'; RS2 = '"R$" #,##0.00'; PCT = '0.0%'

def J(nome):
    return json.loads((DADOS / nome).read_text(encoding='utf-8'))

etapas = J('etapas.json')['etapas']
itens = J('itens.json')['itens']
cots = J('cotacoes.json')['cotacoes']

wb = Workbook()
fina = Side(style='thin', color='C7D8EC')
BORDA = Border(bottom=fina)

def cabecalho(ws, linha, titulos, larguras):
    for i, (t, w) in enumerate(zip(titulos, larguras), start=1):
        c = ws.cell(row=linha, column=i, value=t)
        c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=AZUL)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[linha].height = 26
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)

def corpo(ws, linha, valores, fmt=None, zebra=True):
    for i, v in enumerate(valores, start=1):
        c = ws.cell(row=linha, column=i, value=v)
        c.font = Font(name='Arial', size=9)
        c.border = BORDA
        c.alignment = Alignment(vertical='top', wrap_text=(i in (2, 3, 4, 11, 12)))
        if zebra and linha % 2 == 0:
            c.fill = PatternFill('solid', fgColor=CINZA)
        if fmt and i in fmt:
            c.number_format = fmt[i]

# ============================ 1 · RESUMO ============================
ws = wb.active; ws.title = 'Resumo'
ws['A1'] = 'RESIDÊNCIA Q22 · L8 — PLANILHA DE CUSTOS E MATERIAIS'
ws['A1'].font = Font(name='Arial', size=15, bold=True, color=AZUL)
ws['A2'] = 'Gerada automaticamente a partir dos dados do painel de obra · NÃO editar à mão'
ws['A2'].font = Font(name='Arial', size=9, italic=True, color='5A6B7D')
ws['A3'] = f"Gerada em {datetime.datetime.now(datetime.timezone.utc).astimezone(datetime.timezone(datetime.timedelta(hours=-3))).strftime('%d/%m/%Y %H:%M')} (BRT)"
ws['A3'].font = Font(name='Arial', size=8, color='5A6B7D')
for col, w in zip('ABCD', (34, 20, 20, 46)):
    ws.column_dimensions[col].width = w

n = len(etapas) + 1
linhas = [
    ('INDICADOR', 'VALOR', '', 'OBSERVAÇÃO'),
    ('Área construída', 183, 'm²', 'térreo 75 + garagem 27,5 + superior 62 + extensão 18'),
    ('Estimativa — faixa baixa', f'=SUM(Etapas!C3:C{n + 1})', '', 'soma das 13 etapas'),
    ('Estimativa — faixa alta', f'=SUM(Etapas!D3:D{n + 1})', '', 'soma das 13 etapas'),
    ('Custo/m² — baixo', '=B7/B6', '', 'referência CUB-SP R8-N: R$ 2.129,86'),
    ('Custo/m² — alto', '=B8/B6', '', 'diferença = sistemas premium (bomba de calor, automação, resinados)'),
    ('Total cotado/contratado', f'=SUM(Cotações!K2:K{max(len(cots) + 1, 2)})', '', 'soma de tudo lançado no site'),
    ('  ├ material', f'=SUMIF(Cotações!C2:C{max(len(cots) + 1, 2)},"<>mão de obra",Cotações!K2:K{max(len(cots) + 1, 2)})', '', ''),
    ('  └ mão de obra', f'=SUMIF(Cotações!C2:C{max(len(cots) + 1, 2)},"mão de obra",Cotações!K2:K{max(len(cots) + 1, 2)})', '', 'empreiteiros/equipes cotados por etapa'),
    ('Custo real informado', f'=SUM(Etapas!I3:I{n + 1})', '', 'digitado na aba Etapas do site'),
    ('Cotações registradas', f'=COUNTA(Cotações!A2:A{max(len(cots) + 1, 2)})', '', 'cada lançamento = 1 commit no repositório'),
    ('Avanço geral (ponderado por custo)', f'=IFERROR(SUMPRODUCT(Etapas!H3:H{n + 1},Etapas!E3:E{n + 1})/SUM(Etapas!E3:E{n + 1}),0)', '', 'ponderado pela média da faixa de cada etapa'),
    ('Etapas concluídas', f'=COUNTIF(Etapas!G3:G{n + 1},"concluída")', f'de {len(etapas)}', ''),
]
for i, ln in enumerate(linhas, start=5):
    for j, v in enumerate(ln, start=1):
        c = ws.cell(row=i, column=j, value=v)
        if i == 5:
            c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=AZUL)
        else:
            c.font = Font(name='Arial', size=10, bold=(j == 2), color=(AZUL if j == 2 else '1B2A3A'))
            c.border = BORDA
            c.alignment = Alignment(wrap_text=(j == 4), vertical='center')
for _c in ('B7','B8','B9','B10','B11','B12','B13','B14'):
    ws[_c].number_format = RS
ws['B16'].number_format = PCT

ws['A19'] = 'SISTEMA CONSTRUTIVO (revisão atual)'
ws['A19'].font = Font(name='Arial', size=11, bold=True, color=AZUL)
sist = [
    ('Paredes estruturais', 'Monolev ~320 m² líquidos — perímetro + estruturais do térreo (divisa da garagem, testada)'),
    ('Divisórias em Lightwall', 'Lightwall 2P 90 mm ~100 m²: todo o pav. superior (conceito mezanino) + caixas do lavabo e do banheiro do térreo — nenhuma delas é estrutural'),
    ('Lajes', 'Vigotas + EPS ~86 m² — SÓ pisos (superior, extensão, técnica). Não há laje de forro.'),
    ('Forro', 'Madeira acompanhando o telhado ~45 m² (mezanino + vão duplo) · gesso+EPS nivelado ~47 m² (demais)'),
    ('Cobertura', '3 treliças galvanizadas a fogo @3,25 m (banzos paralelos h=0,40, degrau triangulado) + terças Ue 100 · ~1,0 t'),
    ('Faixa NE (1,20)', 'Da fachada ao fundo: portinha (0,80) · Luz 1 = jardim de inverno c/ parede verde e janelão 3,50 (4,10) · PAREDE CEGA · Luz 2 = varal (4,10) · lavanderia + área técnica acima (4,00, único trecho coberto até a divisa)'),
    ('Fundação', 'Radier ~103 m²'),
    ('Muro de divisa', 'Pré-moldado (placa de concreto) 210 m² — 2 laterais (30,00 m cada, SO/viela + NE/vizinho) + fundo (10,00 m, SE) · h=3,00 m · frente (NO) sem muro (condomínio fechado)'),
    ('Térreo — faixa SO', 'Lavabo 1,50 × 1,50 no canto da garagem c/ o corredor (janela p/ o corredor) · despensa (1,50 × 1,10) + bancada + banheiro completo (1,50 × 2,50, sem box) no fundo, porta ÚNICA p/ a área gourmet'),
]
for i, (a, b) in enumerate(sist, start=20):
    ws.cell(row=i, column=1, value=a).font = Font(name='Arial', size=9, bold=True, color=AZUL2)
    c = ws.cell(row=i, column=2, value=b); c.font = Font(name='Arial', size=9)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[i].height = 24

ws['A29'] = 'COMO USAR'
ws['A29'].font = Font(name='Arial', size=11, bold=True, color=AZUL)
ws['A30'] = ('Esta planilha é um EXPORT do painel de obra e é regerada a cada cotação lançada no site. '
             'Edições feitas aqui serão perdidas na próxima geração — lance tudo pelo site '
             '(pvizioli.github.io/projeto_casa_pedro) para que vire commit e volte para cá.')
ws['A30'].font = Font(name='Arial', size=9, italic=True, color=AMBAR)
ws.merge_cells('A30:D32'); ws['A30'].alignment = Alignment(wrap_text=True, vertical='top')

# ============================ 2 · ETAPAS ============================
ws = wb.create_sheet('Etapas')
ws['A1'] = 'ETAPAS DA OBRA — 13 FASES'
ws['A1'].font = Font(name='Arial', size=13, bold=True, color=AZUL)
cabecalho(ws, 2, ['#', 'Etapa', 'Estim. baixa', 'Estim. alta', 'Média', '% do total',
                  'Status', 'Avanço', 'Custo real', 'Cotado material', 'Cotado mão de obra',
                  'Cotado total', 'Δ vs. alta', 'Observação'],
          [5, 32, 12, 12, 12, 8, 12, 8, 12, 14, 15, 13, 12, 50])
nc = max(len(cots) + 1, 2)
STATUS = {'nao_iniciada': 'não iniciada', 'em_andamento': 'em andamento', 'concluida': 'concluída'}
for i, e in enumerate(etapas):
    r = 3 + i
    corpo(ws, r, [e['id'], e['nome'], e['estimativa_baixa'], e['estimativa_alta'],
                  f'=(C{r}+D{r})/2', f'=IFERROR(E{r}/$E${n + 2},0)', STATUS[e['status']],
                  e['progresso'] / 100, e['custo_real'],
                  f'=SUMIFS(Cotações!$K$2:$K${nc},Cotações!$G$2:$G${nc},$A{r},Cotações!$C$2:$C${nc},"<>mão de obra")',
                  f'=SUMIFS(Cotações!$K$2:$K${nc},Cotações!$G$2:$G${nc},$A{r},Cotações!$C$2:$C${nc},"mão de obra")',
                  f'=J{r}+K{r}',
                  f'=IF(L{r}=0,"",D{r}-L{r})', e['obs']],
          fmt={3: RS, 4: RS, 5: RS, 6: PCT, 8: PCT, 9: RS, 10: RS, 11: RS, 12: RS, 13: RS})
r = n + 2
ws.cell(row=r, column=2, value='TOTAL').font = Font(name='Arial', size=10, bold=True, color=AZUL)
for col, f in ((3, f'=SUM(C3:C{r - 1})'), (4, f'=SUM(D3:D{r - 1})'), (5, f'=SUM(E3:E{r - 1})'),
               (9, f'=SUM(I3:I{r - 1})'), (10, f'=SUM(J3:J{r - 1})'), (11, f'=SUM(K3:K{r - 1})'),
               (12, f'=SUM(L3:L{r - 1})')):
    c = ws.cell(row=r, column=col, value=f)
    c.font = Font(name='Arial', size=10, bold=True, color=AZUL); c.number_format = RS
    c.border = Border(top=Side(style='thin', color=AZUL))
ws.cell(row=r, column=6, value='100%').font = Font(name='Arial', size=10, bold=True, color=AZUL)

# ============================ 3 · ITENS ============================
ws = wb.create_sheet('Itens')
ws['A1'] = 'CATÁLOGO DE ITENS COTÁVEIS'
ws['A1'].font = Font(name='Arial', size=13, bold=True, color=AZUL)
ws['A2'] = 'Base do formulário de cotação do site. Quantidades = memória de cálculo do projeto.'
ws['A2'].font = Font(name='Arial', size=8, italic=True, color='5A6B7D')
cabecalho(ws, 3, ['ID', 'Categoria', 'Tipo', 'Etapa', 'Item', 'Especificação', 'Un', 'Qtde',
                  'Nº cotações', 'Melhor preço unit.', 'Melhor fornecedor', 'Melhor total', 'Observação'],
          [7, 17, 12, 6, 40, 42, 6, 9, 10, 14, 22, 13, 46])
for i, it in enumerate(itens):
    r = 4 + i
    corpo(ws, r, [it['id'], it['categoria'], TIPO.get(it.get('tipo'), 'material'), it.get('etapa'),
                  it['item'], it.get('especificacao') or '',
                  it.get('unidade') or '', it.get('quantidade'),
                  f'=COUNTIFS(Cotações!$E$2:$E${nc},$A{r})',
                  f'=IF(I{r}=0,"—",_xlfn.MINIFS(Cotações!$J$2:$J${nc},Cotações!$E$2:$E${nc},$A{r}))',
                  f'=IF(I{r}=0,"—",IFERROR(INDEX(Cotações!$H$2:$H${nc},MATCH(1,INDEX((Cotações!$E$2:$E${nc}=$A{r})*(Cotações!$J$2:$J${nc}=J{r}),0),0)),""))',
                  f'=IF(I{r}=0,"—",IFERROR(J{r}*H{r},""))',
                  it.get('obs') or ''],
          fmt={8: '#,##0.00', 10: RS2, 12: RS})

# ============================ 4 · COTAÇÕES ============================
ws = wb.create_sheet('Cotações')
cabecalho(ws, 1, ['ID', 'Data', 'Tipo', 'Categoria', 'Item ID', 'Item', 'Etapa',
                  'Fornecedor / empreiteiro', 'Qtde', 'Preço unit.', 'Total', 'Melhor?', 'Observação'],
          [7, 11, 12, 17, 8, 38, 7, 26, 9, 13, 13, 9, 42])
if cots:
    for i, c_ in enumerate(cots):
        r = 2 + i
        corpo(ws, r, [c_['id'], c_.get('data'), TIPO.get(c_.get('tipo'), 'material'), c_.get('categoria'),
                      c_.get('item_id'), c_.get('item'), c_.get('etapa'), c_.get('fornecedor'),
                      c_.get('quantidade'), c_.get('preco_unit'),
                      f'=I{r}*J{r}',
                      f'=IF(J{r}=_xlfn.MINIFS($J$2:$J${nc},$E$2:$E${nc},$E{r}),"SIM","")',
                      c_.get('obs') or ''],
              fmt={9: '#,##0.00', 10: RS2, 11: RS})
else:
    ws['A3'] = 'Nenhuma cotação lançada ainda — use o formulário do site.'
    ws['A3'].font = Font(name='Arial', size=9, italic=True, color=AMBAR)
    ws.merge_cells('A3:M3')

# ============================ 5 · COMPARATIVO ============================
ws = wb.create_sheet('Comparativo')
ws['A1'] = 'COTADO × ESTIMADO POR ETAPA'
ws['A1'].font = Font(name='Arial', size=13, bold=True, color=AZUL)
cabecalho(ws, 2, ['#', 'Etapa', 'Estim. baixa', 'Estim. alta', 'Real / cotado', 'Δ vs. alta', 'Situação'],
          [5, 36, 14, 14, 14, 14, 26])
for i, e in enumerate(etapas):
    r = 3 + i; er = 3 + i
    corpo(ws, r, [e['id'], e['nome'], f'=Etapas!C{er}', f'=Etapas!D{er}',
                  f'=IF(Etapas!I{er}>0,Etapas!I{er},Etapas!L{er})',
                  f'=IF(E{r}=0,"",D{r}-E{r})',
                  f'=IF(E{r}=0,"sem cotação",IF(E{r}>D{r},"ACIMA da faixa",IF(E{r}<C{r},"abaixo da faixa","dentro da faixa")))'],
          fmt={3: RS, 4: RS, 5: RS, 6: RS})

# ============================ 6 · REVISÕES ============================
ws = wb.create_sheet('Revisões')
ws['A1'] = 'HISTÓRICO DE REVISÕES DO PROJETO'
ws['A1'].font = Font(name='Arial', size=13, bold=True, color=AZUL)
cabecalho(ws, 2, ['Data', 'Alteração', 'Impacto'], [12, 62, 62])
revs = [
    ('13/07/2026', 'Versão inicial da planilha (13 etapas, 116 itens)', 'Faixa R$ 498–764 mil'),
    ('13/07/2026', 'Orientação corrigida: frente NOROESTE (não norte)', 'Fachada frontal recebe sol da tarde — verificar conforto'),
    ('13/07/2026', 'Janelão da escada c/ peitoril 1,00 m (2,2 × 4,70)', 'Vidro 12 → 10,5 m²'),
    ('13/07/2026', 'Vão da garagem em altura total (chão → laje), viga na testada', 'Perfil da viga a validar (altura útil limitada)'),
    ('15/07/2026', 'Esclarecido: lajes = SÓ pisos. Forro nivelado gesso+EPS incluído', '+47 m² de forro (antes ausente do orçamento)'),
    ('15/07/2026', 'Forro de MADEIRA no mezanino/vão duplo acompanhando o telhado', '+45 m² · R$ 8–16 mil (item premium)'),
    ('15/07/2026', 'Divisórias do superior migradas de Monolev p/ Lightwall 2P 90', 'Monolev 421 → 340 m² · Etapa 4: R$ 120–170k → R$ 110–156k'),
    ('16/07/2026', 'Faixa NE: área de luz até a fachada, portinha recuada 0,80', 'Telha 105 → 98 m² · estrutura 100 → 95 m²'),
    ('16/07/2026', 'Parede cega entre Luz 1 e Luz 2 · Luz 1 = jardim de inverno', '+14 m² parede verde · +12 m² janelão · impermeabilizar divisa'),
    ('16/07/2026', 'Cobertura: pórticos de viga cheia → 3 TRELIÇAS @3,25 m', 'Aço 1,4 t → ~1,0 t · degrau do lanternim triangulado'),
    ('16/07/2026', 'Terças: 6 → 4 linhas e Ue 150 → Ue 100', '−170 kg · σ 61 MPa (24% de fy), flecha L/483'),
    ('16/07/2026', 'Custo de projeto/engenheiro/arquiteto detalhado (8 itens)', 'Etapa 1: R$ 20–35k → R$ 34–78k · total R$ 502–793k'),
    ('16/07/2026', 'Mão de obra cotável por etapa (13 itens) + aba Fornecedores', 'Material e mão de obra separados por etapa e por fornecedor'),
    ('16/07/2026', 'Térreo: banheiro sai da parede da garagem → banheiro + lavabo no fundo da cozinha (faixa SO)', 'Cozinha 36 → 29 m² · +1 lavabo · janelas p/ o acesso lateral · porta do lavabo p/ a varanda'),
    ('16/07/2026', 'Porta da lavanderia a 0,90 da parede do fundo', 'Máquina sob a janela do fundo + nicho de geladeira no canto da cozinha'),
    ('16/07/2026', 'Lavanderia/técnica 2,50 → 4,00 m; Luz 1 e Luz 2 4,85 → 4,10 m cada', 'Faixa de luz 9,70 → 8,20 m · laje 68 → 70 m² · telha 98 → 100 m² · parede verde 14 → 12 m²'),
    ('16/07/2026', 'Banheiro e lavabo do térreo unificados em um só (1,50 × 2,50)', 'Cozinha recupera ~1,9 m² · porta única p/ a área gourmet'),
    ('16/07/2026', 'Lavabo 1,50 × 1,50 no canto da garagem c/ o corredor (janela p/ o corredor)', 'Térreo volta a ter banheiro acessível por dentro'),
    ('16/07/2026', 'Divisórias do térreo (lavabo e banheiro) em Lightwall, não Monolev', 'Monolev 340 → 320 m² · Lightwall 80 → 100 m² · esgoto de 100 mm obrigatoriamente pelo radier'),
    ('16/07/2026', 'Superior: copa desce 1 m; corredor avança 1 m na suíte; porta da suíte na lateral (closet recuado)', 'Quarto 3,70 → 4,70 m (10 → 11,8 m²) · suíte 4,50 → 3,50 m · closet passante 1,20 × 2,50 · suíte sem janela na viela (luz pelo vidro da extensão)'),
    ('16/07/2026', 'Corredor reto, porta da suíte de frente no fim do corredor (desfeita a porta lateral); parede do closet removida (ficava dupla c/ a do corredor)', 'Closet passa de 1,20 → 1,60 m de largura; nicho natural atrás da parede do corredor'),
    ('16/07/2026', 'Cobogó em parede flutuante em frente ao janelão da fachada NO (h=3,50, afastado 1,00 m); porta de acesso à Luz 1 removida (a incluir depois); película espelhada restrita ao piso superior', 'Câmara de ar ventilada evita transferência de calor para o vidro · térreo do janelão fica com vidro transparente'),
    ('16/07/2026', 'Muro de divisa pré-moldado nos 2 lados + fundo (210 m², h=3,00) — frente sem muro (condomínio fechado)', 'Etapa 2 (fundação): R$ 45k → R$ 50,5-78,5k · total do projeto R$ 502-793k → R$ 533,5-845,5k'),
    ('16/07/2026', 'Cobogó da fachada NO ajustado para h=3,00 (era 3,50) — mesma altura do muro de divisa', 'Uniformiza a altura dos elementos verticais externos da casa'),
    ('16/07/2026', 'Cobogó estendido até encostar no muro de divisa NE (antes parava ~0,80 m antes)', 'Vira estrutura contínua com o muro — 1 ponta apoiada no mourão do muro, só a ponta livre precisa de mourão próprio'),
    ('19/07/2026', 'Cantoneira de PVC c/ abas de tela especificada para TODOS os cantos externos e vãos (321 m)', 'Padroniza o arremate em Monolev e Lightwall · alumínio descartado em massa cimentícia (corrosão alcalina) · ~R$ 1,6-2,6 mil, dentro da etapa 4'),
    ('19/07/2026', 'Auditoria de orientação das folhas: planta e elevações conferidas e coerentes; marcadores NE/SO adicionados nas elevações', 'A elevação da rua é espelhada por convenção (NE à esquerda, SO à direita) — agora indicado na própria folha'),
    ('19/07/2026', 'Janelão da escada corrigido para 3,20 x 4,70 m (havia 3 valores divergentes: 3,18 na planta, 2,40 na elevação, 2,20 no catálogo)', 'Vidro 10,5 -> 15 m² · impacto na etapa 7 (esquadrias)'),
]
for i, (d, a, imp) in enumerate(revs):
    corpo(ws, 3 + i, [d, a, imp])

# ============================ 7 · FORNECEDORES ============================
ws = wb.create_sheet('Fornecedores')
ws['A1'] = 'CONSOLIDADO POR FORNECEDOR / EMPREITEIRO'
ws['A1'].font = Font(name='Arial', size=13, bold=True, color=AZUL)
ws['A2'] = 'Uma linha por fornecedor cotado no site — quem fornece o quê, em qual etapa e por quanto.'
ws['A2'].font = Font(name='Arial', size=8, italic=True, color='5A6B7D')
cabecalho(ws, 3, ['Fornecedor / empreiteiro', 'Tipo', 'Nº de cotações', 'Total cotado',
                  'Etapas envolvidas', 'Itens'], [30, 14, 12, 15, 20, 74])
forns = {}
for c_ in cots:
    f = (c_.get('fornecedor') or '—').strip()
    d_ = forns.setdefault(f, {'n': 0, 'total': 0.0, 'etapas': set(), 'itens': [], 'tipos': set()})
    d_['n'] += 1
    d_['total'] += (c_.get('quantidade') or 0) * (c_.get('preco_unit') or 0)
    if c_.get('etapa'): d_['etapas'].add(c_['etapa'])
    d_['itens'].append(c_.get('item') or '')
    d_['tipos'].add(TIPO.get(c_.get('tipo'), 'material'))
if forns:
    for i, (f, v) in enumerate(sorted(forns.items(), key=lambda x: -x[1]['total'])):
        r = 4 + i
        corpo(ws, r, [f, ' + '.join(sorted(v['tipos'])), v['n'],
                      f'=SUMIF(Cotações!$H$2:$H${nc},$A{r},Cotações!$K$2:$K${nc})',
                      ', '.join(str(x) for x in sorted(v['etapas'])),
                      ' · '.join(dict.fromkeys(v['itens']))],
              fmt={4: RS})
    r = 4 + len(forns)
    ws.cell(row=r, column=1, value='TOTAL').font = Font(name='Arial', size=10, bold=True, color=AZUL)
    c = ws.cell(row=r, column=4, value=f'=SUM(D4:D{r - 1})')
    c.font = Font(name='Arial', size=10, bold=True, color=AZUL); c.number_format = RS
    c.border = Border(top=Side(style='thin', color=AZUL))
else:
    ws['A4'] = 'Nenhum fornecedor cotado ainda. Cada cotação lançada no site cria/alimenta uma linha aqui.'
    ws['A4'].font = Font(name='Arial', size=9, italic=True, color=AMBAR)
    ws.merge_cells('A4:F4')

wb.save(SAIDA)
print(f'OK · {SAIDA.name} · {len(etapas)} etapas · {len(itens)} itens · {len(cots)} cotações')
